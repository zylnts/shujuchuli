import openpyxl
import os

# 获取桩号的函数
def get_zh(z_qz, z_bh):
    if z_bh and z_qz:
        try:
            z_bh_value = float(z_bh)
            if z_bh_value.is_integer():
                z_bh_value = int(z_bh_value)
        except ValueError:
            return None
        
        z_bh2 = z_bh_value % 1000 if z_bh_value >= 1000 else z_bh_value
        return f"{z_qz}{z_bh_value // 1000}+{z_bh2:.3f}" if z_bh_value >= 1000 else f"{z_qz}0+{z_bh2:.3f}"
    return None

# 创建最终数据的函数
def create_data_e(mf_row, mo_row, 原始数据, 输出数据):
    m_zlxx = 原始数据.cell(row=mf_row, column=3).value
    mark = ["<#0>", "<#1>", "<#2>", "<#3>", "<#4>", "<#5>", "<#6>", "<#7>"]
    
    # 将原始数据复制到输出数据表中
    输出数据.cell(row=mo_row, column=4).value = 原始数据.cell(row=mf_row, column=2).value
    mo_row += 1
    输出数据.cell(row=mo_row, column=4).value = 原始数据.cell(row=mf_row, column=1).value
    mo_row += 1
    输出数据.cell(row=mo_row, column=4).value = 原始数据.cell(row=mf_row, column=5).value
    mo_row += 1
    
    # 设置输出数据的头部
    headers = ["COORDINATES", "0", "FLOW DIRECTION", "0", "DATUM", "0", "RADIUS TYPE", "0",
               "DIVIDE X-Section", "0", "SECTION ID", 原始数据.cell(row=mf_row, column=10).value, "INTERPOLATED", "0",
               "ANGLE", "0.00   0", "RESISTANCE NUMBERS", "2  1     1.000     1.000     1.000    1.000    1.000"]
    
    for header in headers:
        输出数据.cell(row=mo_row, column=4).value = header
        mo_row += 1
    
    # 处理每个点的数据
    for i in range(6):
        m_str_tmp = 原始数据.cell(row=mf_row, column=11 + i).value
        tmp1, tmp2 = m_str_tmp.split('\n')[:2]  # 使用换行符分割数据
        输出数据.cell(row=mo_row, column=4).value = tmp1
        输出数据.cell(row=mo_row, column=5).value = tmp2
        输出数据.cell(row=mo_row, column=6).value = m_zlxx
        输出数据.cell(row=mo_row, column=7).value = mark[i] if i in [1, 2, 4] else mark[0]
        mo_row += 1
    
    输出数据.cell(row=mo_row, column=4).value = "*****************************"
    mo_row += 1

# 创建中间数据的函数
def create_data_m(原始数据, 输出数据):
    f_row = 2  # 起始行号
    o_row = 1
    o_row2 = 1
    
    # 处理原始数据，生成中间数据
    while 原始数据.cell(row=f_row, column=1).value is not None:
        print(f"Processing row {f_row}")  # 添加调试语句
        
        原始数据.cell(row=f_row, column=10).value = get_zh(原始数据.cell(row=f_row, column=4).value, 原始数据.cell(row=f_row, column=5).value)  # D0
        输出数据.cell(row=o_row, column=1).value = 原始数据.cell(row=f_row, column=10).value
        o_row += 1
        
        m_num = 0
        m_num += 10
        原始数据.cell(row=f_row, column=11).value = f"{m_num:.3f}\n{原始数据.cell(row=f_row, column=6).value:.3f}"  # D1
        输出数据.cell(row=o_row, column=1).value = f"{m_num:.3f}"
        输出数据.cell(row=o_row, column=2).value = f"{原始数据.cell(row=f_row, column=6).value:.3f}"
        o_row += 1
        
        m_num += (原始数据.cell(row=f_row, column=6).value - 原始数据.cell(row=f_row, column=8).value) * 原始数据.cell(row=f_row, column=9).value
        原始数据.cell(row=f_row, column=12).value = f"{m_num:.3f}\n{原始数据.cell(row=f_row, column=6).value:.3f}"  # D2
        输出数据.cell(row=o_row, column=1).value = f"{m_num:.3f}"
        输出数据.cell(row=o_row, column=2).value = f"{原始数据.cell(row=f_row, column=6).value:.3f}"
        o_row += 1
        
        m_num += 原始数据.cell(row=f_row, column=7).value
        原始数据.cell(row=f_row, column=13).value = f"{m_num:.3f}\n{原始数据.cell(row=f_row, column=8).value:.3f}"  # D3
        输出数据.cell(row=o_row, column=1).value = f"{m_num:.3f}"
        输出数据.cell(row=o_row, column=2).value = f"{原始数据.cell(row=f_row, column=8).value:.3f}"
        o_row += 1
        
        m_num += (原始数据.cell(row=f_row, column=6).value - 原始数据.cell(row=f_row, column=8).value) * 原始数据.cell(row=f_row, column=9).value
        原始数据.cell(row=f_row, column=14).value = f"{m_num:.3f}\n{原始数据.cell(row=f_row, column=8).value:.3f}"  # D4
        输出数据.cell(row=o_row, column=1).value = f"{m_num:.3f}"
        输出数据.cell(row=o_row, column=2).value = f"{原始数据.cell(row=f_row, column=8).value:.3f}"
        o_row += 1
        
        m_num += 10
        原始数据.cell(row=f_row, column=15).value = f"{m_num:.3f}\n{原始数据.cell(row=f_row, column=6).value:.3f}"  # D5
        输出数据.cell(row=o_row, column=1).value = f"{m_num:.3f}"
        输出数据.cell(row=o_row, column=2).value = f"{原始数据.cell(row=f_row, column=6).value:.3f}"
        o_row += 1
        
        原始数据.cell(row=f_row, column=16).value = f"{m_num:.3f}\n{原始数据.cell(row=f_row, column=6).value:.3f}"  # D6
        输出数据.cell(row=o_row, column=1).value = f"{m_num:.3f}"
        输出数据.cell(row=o_row, column=2).value = f"{原始数据.cell(row=f_row, column=6).value:.3f}"
        o_row += 1
        
        create_data_e(f_row, o_row2, 原始数据, 输出数据)
        
        f_row += 1
    
    save_xyz(输出数据)

# 保存 XYZ 文件的函数
def save_xyz(ws):
    xyz_save_file = input("请输入要保存的txt文件路径：")
    
    if xyz_save_file:
        if os.path.exists(xyz_save_file):
            overwrite = input("输入的文件已存在，是否覆盖？(Y/N): ")
            if overwrite.lower() != 'y':
                return save_xyz(ws)

        with open(xyz_save_file, 'w') as f:
            for row in range(1, ws.max_row + 1):
                o_tmp = ws.cell(row=row, column=4).value
                if o_tmp:
                    f.write(f"{o_tmp}\t{ws.cell(row=row, column=5).value}\t{ws.cell(row=row, column=6).value}\t{ws.cell(row=row, column=7).value}\n")

        print("文件已保存。")

# 打开 Excel 工作簿
wb = openpyxl.load_workbook('C:/Users/Zhangyl/Desktop/1.xlsx')
# 选择活动工作表
ws_original = wb.active
ws_output = wb.create_sheet("输出数据")  # 创建输出数据的工作表

# 调用函数创建中间数据并保存为 txt 文件
create_data_m(ws_original, ws_output)
